"""
Read 数据.xlsx (same folder) and emit import_data.sql for MySQL.
- Maps Excel dicharge_doctor_id -> discharge_doctor_id
- Maps DOCTORS.speciality (Excel) -> DOCTORS.speciality column
- TREATMENT_LOG.to_dept_id (Excel) matches SQL
- Skips blank tail rows in DISCHARGES
- Fills missing TREATMENT_LOG.log_id
- Fixes: non-F patient in Obstetrics & Gynecology (dept 4) -> dept 1 (Internal Medicine)
- If admission has no TREATMENT_LOG rows, discharge_dept_id is forced to admission_dept_id (matches DB trigger)
- INSERT order: ADMISSIONS -> TREATMENT_LOG -> DISCHARGES (so discharge trigger sees logs)
- Reports fixes to STDERR and import_report.txt

Requires: pip install openpyxl pypinyin
"""

from __future__ import annotations

import glob
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

OBGYN_DEPT_ID = 4
FIX_DEPT_ID = 1  # Internal Medicine

_CJK_RE = re.compile(r"[\u4e00-\u9fff]")
_PINYIN = None  # (Style, lazy_pinyin) set in main()


def chinese_name_to_pinyin(val):
    """Convert Chinese personal names to ASCII pinyin (TitleCase syllables). Non-CJK unchanged."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.upper() == "NULL":
        return val
    if not _CJK_RE.search(s):
        return val
    Style, lazy_pinyin = _PINYIN
    parts = lazy_pinyin(s, style=Style.NORMAL, errors="ignore")
    out = "".join((p or "").capitalize() for p in parts if p)
    return out or s


def find_xlsx() -> Path:
    base = Path(r"c:\Users\Lenovo\Desktop\新建文件夹")
    cands = list(base.glob("*.xlsx"))
    cands = [p for p in cands if not p.name.startswith("~$")]
    if not cands:
        cands = [
            p
            for p in glob.glob(str(base / "**" / "*.xlsx"), recursive=True)
            if "~$" not in p
        ]
        cands = [Path(p) for p in cands]
    if not cands:
        raise FileNotFoundError("No .xlsx found in 新建文件夹")
    return max(cands, key=lambda p: p.stat().st_mtime)


def esc(val) -> str:
    if val is None:
        return "NULL"
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, Decimal):
        return str(val)
    if isinstance(val, datetime):
        return esc(val.date())
    if isinstance(val, date):
        return f"'{val.isoformat()}'"
    s = str(val).strip()
    if s.upper() == "NULL" or s == "":
        return "NULL"
    s = s.replace("\\", "\\\\").replace("'", "''")
    s = re.sub(r"[\x00-\x1f]", " ", s)
    return f"'{s}'"


def main() -> None:
    global _PINYIN
    try:
        import openpyxl
    except ImportError:
        print("Please: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    try:
        from pypinyin import Style, lazy_pinyin

        _PINYIN = (Style, lazy_pinyin)
    except ImportError:
        print("Please: pip install pypinyin", file=sys.stderr)
        sys.exit(1)

    xlsx = find_xlsx()
    out_sql = Path(__file__).resolve().parent / "import_data.sql"
    out_report = Path(__file__).resolve().parent / "import_report.txt"
    lines: list[str] = []
    report: list[str] = []

    wb = openpyxl.load_workbook(xlsx, data_only=True)

    # --- Reference tables (order matters) ---
    def rows(sheet: str, max_col: int | None = None):
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is None:
                continue
            row = []
            mc = max_col or ws.max_column
            for c in range(1, mc + 1):
                row.append(ws.cell(r, c).value)
            yield row

    lines.append("USE csc3170_hospital;")
    lines.append("SET FOREIGN_KEY_CHECKS=0;")
    lines.append("SET UNIQUE_CHECKS=0;")
    lines.append("SET NAMES utf8mb4;")
    lines.append(
        "TRUNCATE TABLE TREATMENT_LOG;"
        " TRUNCATE TABLE DISCHARGES;"
        " TRUNCATE TABLE ADMISSIONS;"
        " TRUNCATE TABLE DOCTORS;"
        " TRUNCATE TABLE PATIENTS;"
        " TRUNCATE TABLE PATIENT_EMERGENCY_CONTACT;"
        " TRUNCATE TABLE DOCTOR_EMERGENCY_CONTACT;"
        " TRUNCATE TABLE DEPARTMENTS;"
        " TRUNCATE TABLE PROVINCES;"
    )

    for name, table, cols in [
        ("PROVINCES", "PROVINCES", ["province_id", "province_name"]),
        (
            "PATIENT_EMERGENCY_CONTACT",
            "PATIENT_EMERGENCY_CONTACT",
            ["pec_id", "first_name", "last_name", "phone"],
        ),
        (
            "DOCTOR_EMERGENCY_CONTACT",
            "DOCTOR_EMERGENCY_CONTACT",
            ["dec_id", "first_name", "last_name", "phone"],
        ),
        ("DEPARTMENTS", "DEPARTMENTS", ["dept_id", "dept_name"]),
    ]:
        buf = []
        ws = wb[name]
        hdr = [ws.cell(1, c).value for c in range(1, len(cols) + 1)]
        assert [h.lower() if isinstance(h, str) else h for h in hdr][: len(cols)] == [
            c.lower() for c in cols
        ], (name, hdr, cols)
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value is None:
                continue
            vals = [ws.cell(r, i + 1).value for i in range(len(cols))]
            if name in ("PATIENT_EMERGENCY_CONTACT", "DOCTOR_EMERGENCY_CONTACT"):
                vals[1] = chinese_name_to_pinyin(vals[1])
                vals[2] = chinese_name_to_pinyin(vals[2])
            buf.append(f"({', '.join(esc(v) for v in vals)})")
        lines.append(
            f"INSERT INTO {table} ({', '.join(cols)}) VALUES\n"
            + ",\n".join(buf)
            + ";"
        )

    # PATIENTS
    ws = wb["PATIENT"]
    pcols = [
        "patient_id",
        "first_name",
        "last_name",
        "gender",
        "birth_date",
        "city",
        "allergies",
        "height",
        "weight",
        "phone",
        "province_id",
        "pec_id",
    ]
    buf = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        row = []
        row.append(ws.cell(r, 2).value)  # patient_id col2
        row.append(chinese_name_to_pinyin(ws.cell(r, 3).value))
        row.append(chinese_name_to_pinyin(ws.cell(r, 4).value))
        row.append(ws.cell(r, 5).value)
        bd = ws.cell(r, 6).value
        if isinstance(bd, datetime):
            bd = bd.date()
        row.append(bd)
        row.append(ws.cell(r, 7).value)
        al = ws.cell(r, 8).value
        if isinstance(al, str) and al.strip().upper() == "NULL":
            al = None
        row.append(al)
        for c in (9, 10):
            v = ws.cell(r, c).value
            if v is not None and not isinstance(v, (int,)):
                v = float(v) if isinstance(v, (float,)) else v
            row.append(v)
        ph = ws.cell(r, 11).value
        if ph is not None:
            if isinstance(ph, float):
                ph = str(int(ph)) if ph == int(ph) else str(ph)
            else:
                ph = str(ph).strip()
        row.append(ph)
        row.append(ws.cell(r, 12).value)
        row.append(ws.cell(r, 13).value)
        buf.append(f"({', '.join(esc(v) for v in row)})")
    lines.append(
        f"INSERT INTO PATIENTS ({', '.join(pcols)}) VALUES\n"
        + ",\n".join(buf)
        + ";"
    )

    gender_by_patient = {}
    ws = wb["PATIENT"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        gender_by_patient[int(ws.cell(r, 2).value)] = ws.cell(r, 5).value

    # DOCTORS (speciality)
    ws = wb["DOCTORS"]
    dcols = [
        "doctor_id",
        "first_name",
        "last_name",
        "speciality",
        "years_of_exp",
        "phone",
        "dept_id",
        "dec_id",
    ]
    buf = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        vals = [ws.cell(r, c).value for c in range(1, 9)]
        ph = vals[5]
        if isinstance(ph, float):
            vals[5] = str(int(ph)) if ph == int(ph) else str(ph)
        y = vals[4]
        if isinstance(y, float):
            vals[4] = int(y)
        vals[1] = chinese_name_to_pinyin(vals[1])
        vals[2] = chinese_name_to_pinyin(vals[2])
        buf.append(f"({', '.join(esc(v) for v in vals)})")
    lines.append(
        f"INSERT INTO DOCTORS ({', '.join(dcols)}) VALUES\n"
        + ",\n".join(buf)
        + ";"
    )

    # ADMISSIONS (fix OBGYN for non-F)
    ws = wb["ADMISSIONS"]
    acols = [
        "admission_id",
        "patient_id",
        "attending_doctor_id",
        "admission_dept_id",
        "admission_date",
        "diagnosis",
        "diag_severity",
    ]
    buf = []
    admission_dept_by_id: dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        aid = int(ws.cell(r, 1).value)
        pid = int(ws.cell(r, 2).value)
        dept = int(ws.cell(r, 4).value)
        g = gender_by_patient.get(pid)
        if dept == OBGYN_DEPT_ID and g not in ("F",):
            report.append(
                f"Row {r} admission_id={aid} patient={pid} gender={g}: "
                f"admission_dept_id {OBGYN_DEPT_ID}-> {FIX_DEPT_ID} (non-F in OB/GYN)"
            )
            dept = FIX_DEPT_ID
        admission_dept_by_id[aid] = dept
        ad = ws.cell(r, 5).value
        if isinstance(ad, datetime):
            ad = ad.date()
        diag = ws.cell(r, 6).value or ""
        sev = ws.cell(r, 7).value
        if isinstance(sev, float):
            sev = int(sev)
        vals = [
            aid,
            pid,
            ws.cell(r, 3).value,
            dept,
            ad,
            diag,
            sev,
        ]
        buf.append(f"({', '.join(esc(v) for v in vals)})")
    lines.append(
        f"INSERT INTO ADMISSIONS ({', '.join(acols)}) VALUES\n"
        + ",\n".join(buf)
        + ";"
    )

    admissions_with_log: set[int] = set()
    tws = wb["TREATMENT_LOG"]
    for r in range(2, tws.max_row + 1):
        if tws.cell(r, 2).value is None and tws.cell(r, 1).value is None:
            continue
        admissions_with_log.add(int(tws.cell(r, 2).value))

    # TREATMENT_LOG (before DISCHARGES so triggers see logs on discharge insert)
    ws = wb["TREATMENT_LOG"]
    tcols = [
        "log_id",
        "admission_id",
        "examining_doctor_id",
        "to_dept_id",
        "exam_result",
        "exam_date",
    ]
    buf = []
    next_id = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None and ws.cell(r, 1).value is None:
            continue
        log_id = ws.cell(r, 1).value
        if log_id is None:
            log_id = next_id
            report.append(f"TREATMENT_LOG row {r}: generated log_id={log_id}")
        else:
            log_id = int(log_id)
        next_id = max(next_id, int(log_id) + 1)
        aid = int(ws.cell(r, 2).value)
        exdoc = ws.cell(r, 3).value
        todept = int(ws.cell(r, 4).value)
        # patient gender for OBGYN
        pid = None
        aw = wb["ADMISSIONS"]
        for ar in range(2, aw.max_row + 1):
            if aw.cell(ar, 1).value == aid:
                pid = int(aw.cell(ar, 2).value)
                break
        g = gender_by_patient.get(pid)
        if todept == OBGYN_DEPT_ID and g not in ("F",):
            report.append(
                f"TREATMENT_LOG row {r} log_id={log_id}: to_dept_id {OBGYN_DEPT_ID}-> {FIX_DEPT_ID}"
            )
            todept = FIX_DEPT_ID
        txt = ws.cell(r, 5).value or ""
        exd = ws.cell(r, 6).value
        if isinstance(exd, datetime):
            exd = exd.date()
        vals = [log_id, aid, exdoc, todept, txt, exd]
        buf.append(f"({', '.join(esc(v) for v in vals)})")
    lines.append(
        f"INSERT INTO TREATMENT_LOG ({', '.join(tcols)}) VALUES\n"
        + ",\n".join(buf)
        + ";"
    )

    # DISCHARGES (skip empty; fix typo column; fix OBGYN; no-log -> discharge dept = admission dept)
    ws = wb["DISCHARGES"]
    dcols = [
        "discharge_id",
        "admission_id",
        "discharge_dept_id",
        "discharge_doctor_id",
        "discharge_date",
        "discharge_status",
    ]
    buf = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue
        did = int(ws.cell(r, 1).value)
        aid = int(ws.cell(r, 2).value)
        # map patient via admissions sheet
        arow = None
        aw = wb["ADMISSIONS"]
        for ar in range(2, aw.max_row + 1):
            if aw.cell(ar, 1).value == aid:
                arow = int(aw.cell(ar, 2).value)
                break
        if arow is None:
            report.append(f"WARNING discharge row {r}: admission_id {aid} not found in ADMISSIONS")
            continue
        g = gender_by_patient.get(arow)
        ddpt = int(ws.cell(r, 3).value)
        if ddpt == OBGYN_DEPT_ID and g not in ("F",):
            report.append(
                f"DISCHARGES row {r} discharge_id={did}: discharge_dept_id {OBGYN_DEPT_ID}-> {FIX_DEPT_ID}"
            )
            ddpt = FIX_DEPT_ID
        adm_dept = admission_dept_by_id.get(aid)
        if adm_dept is not None and aid not in admissions_with_log and ddpt != adm_dept:
            report.append(
                f"DISCHARGES row {r} discharge_id={did} admission_id={aid}: "
                f"no TREATMENT_LOG — discharge_dept_id {ddpt} -> {adm_dept} (must match admission_dept)"
            )
            ddpt = adm_dept
        doc = ws.cell(r, 4).value  # dicharge in Excel
        ddate = ws.cell(r, 5).value
        if isinstance(ddate, datetime):
            ddate = ddate.date()
        st = ws.cell(r, 6).value
        vals = [did, aid, ddpt, doc, ddate, st]
        buf.append(f"({', '.join(esc(v) for v in vals)})")
    lines.append(
        f"INSERT INTO DISCHARGES ({', '.join(dcols)}) VALUES\n"
        + ",\n".join(buf)
        + ";"
    )

    lines.append("SET FOREIGN_KEY_CHECKS=1;")
    lines.append("SET UNIQUE_CHECKS=1;")

    out_sql.write_text("\n".join(lines) + "\n", encoding="utf-8")
    out_report.write_text(
        "\n".join(
            [
                f"Source: {xlsx}",
                f"Rows / fixes: {len(report)}",
                "",
                *report,
            ]
        ),
        encoding="utf-8",
    )
    print(f"Wrote {out_sql}", file=sys.stderr)
    print(f"Wrote {out_report}", file=sys.stderr)


if __name__ == "__main__":
    main()
